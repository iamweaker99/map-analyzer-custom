"""Build motor-plan Excel (per-pattern MPA/MM/SC) from prototype_sequence_motor outputs.

Usage: python build_motor_xlsx.py <patterns.json> <seq.json> <out.xlsx>

- patterns.json: output of `prototype_sequence_motor --patterns --R --exp`
  (all patterns incl. singletons; time_ms = start time of first note in ms)
- seq.json: output of `prototype_sequence_motor --json --R --exp`
  (per multi-note pattern: time (mm:ss:mmm), notes, mpa, mm, sc)

Merge key = (rounded start ms, note count) — mirrors the prototype's
format_time rounding (Rust round = half away from zero, NOT Python banker's).
Singletons have no motor metrics (skipped by analyze_patterns) -> blank cells.
"""

import json
import math
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["Time", "Type", "Notes", "Range", "Snap", "MPA", "MM", "SC"]
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def rust_round(x: float) -> int:
    """Match Rust f64::round() (half away from zero), used by format_time."""
    return math.floor(x + 0.5)


def fmt_time(ms: float) -> str:
    """mm:ss:mmm, same zero-padding as the prototype's format_time."""
    ms = rust_round(ms)
    mins = ms // 60000
    secs = (ms // 1000) % 60
    millis = ms % 1000
    return f"{mins:02d}:{secs:02d}:{millis:03d}"


def parse_time(s: str) -> int:
    m, ss, ms = s.split(":")
    return int(m) * 60000 + int(ss) * 1000 + int(ms)


def main(patterns_path: str, seq_path: str, out_path: str) -> None:
    with open(patterns_path, encoding="utf-8") as f:
        patterns = json.load(f)["patterns"]
    with open(seq_path, encoding="utf-8") as f:
        timeline = json.load(f)["timeline"]

    seq_map: dict[tuple[int, int], tuple[float, float, float]] = {}
    for entry in timeline:
        key = (parse_time(entry["time"]), entry["notes"])
        if key in seq_map:
            print(f"WARNING: duplicate seq entry {key}")
        seq_map[key] = (entry["mpa"], entry["mm"], entry["sc"])

    multi_count = sum(1 for p in patterns if p["notes"] >= 2)
    if multi_count != len(timeline):
        print(
            f"WARNING: multi-note patterns ({multi_count}) != "
            f"timeline entries ({len(timeline)})"
        )

    rows = []
    misses = 0
    for p in patterns:
        key = (rust_round(p["time_ms"]), p["notes"])
        vals = seq_map.get(key)
        if p["notes"] >= 2 and vals is None:
            misses += 1
            print(f"  MISS: {key} {p['type']}")
        mpa, mm, sc = vals if vals else (None, None, None)
        if mpa is not None:
            mpa, mm, sc = round(mpa, 3), round(mm, 3), round(sc, 3)
        rows.append(
            [fmt_time(p["time_ms"]), p["type"], p["notes"], p["range"], p["snap"], mpa, mm, sc]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Patterns"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(h) + 4)

    wb.save(out_path)
    print(
        f"wrote {out_path}: {len(rows)} rows "
        f"(multi-note {multi_count}, timeline {len(timeline)}, misses {misses})"
    )


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])
